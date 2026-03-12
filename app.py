import os
import re
import shutil
import subprocess
import tempfile
from io import BytesIO

import streamlit as st
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="PO PDF to Excel", page_icon="📄", layout="wide")
st.title("📄 Purchase Order PDFs → One Excel")
st.caption("Upload multiple PO PDFs, extract line items, and download one combined XLSX.")

ITEM_CODE_RE = re.compile(r"^[A-Z0-9][A-Z0-9\-]{3,}$")
NUM_RE = re.compile(r"^\d+(?:\.\d+)?$")


def extract_text_from_pdf(file) -> str:
    """
    Prefer Poppler pdftotext (better layout/text recovery on PO PDFs),
    then fallback to pypdf if pdftotext is unavailable/fails.
    """
    # Streamlit upload file-like object can be read multiple times if we reset pointer
    file.seek(0)

    pdftotext_bin = shutil.which("pdftotext")
    if pdftotext_bin:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_pdf:
            tmp_pdf.write(file.read())
            pdf_path = tmp_pdf.name

        txt_path = pdf_path + ".txt"
        try:
            # -layout preserves columns better for table-like PDFs
            subprocess.run([pdftotext_bin, "-layout", pdf_path, txt_path], check=True)
            with open(txt_path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        except Exception:
            # fall through to pypdf fallback
            pass
        finally:
            for p in (pdf_path, txt_path):
                if os.path.exists(p):
                    os.remove(p)

    # Fallback: pure-python extraction
    file.seek(0)
    reader = PdfReader(file)
    parts = []
    for page in reader.pages:
        parts.append(page.extract_text() or "")
    return "\n".join(parts)


def parse_po_number(text: str) -> str:
    # Many Computer Mania docs carry an 8-digit PO/doc number near date + page footer.
    m = re.search(r"\b\d{2}/\d{2}/\d{4}\s+(\d{8})\s+Page\s*:\s*\d+", text, re.IGNORECASE)
    if m:
        return m.group(1)

    # Fallbacks for other layouts
    m = re.search(r"(\d{10})\s*\n\s*PURCHASE ORDER NUMBER", text, re.IGNORECASE)
    if m:
        return m.group(1)
    m = re.search(r"PURCHASE ORDER NUMBER\s*\n\s*(\d{6,12})", text, re.IGNORECASE)
    if m:
        return m.group(1)
    m = re.search(r"\b\d{8}\b", text)
    if m:
        return m.group(0)
    m = re.search(r"\b\d{10}\b", text)
    return m.group(0) if m else ""


def parse_po_date(text: str) -> str:
    m = re.search(r"PO Date:\s*([0-3]?\d/[01]?\d/\d{4})", text, re.IGNORECASE)
    if m:
        return m.group(1)
    m = re.search(r"\b([0-3]?\d/[01]?\d/\d{4})\b", text)
    return m.group(1) if m else ""


def parse_rows(text: str):
    rows = []
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    def is_num(s: str) -> bool:
        return bool(NUM_RE.match(s))

    def try_map_numbers(nums):
        """Try common column orders and pick best consistency score."""
        candidates = []
        if len(nums) < 5:
            return None

        # Format A: qty, price, total_excl, tax, total_incl
        a = {
            "quantity": nums[0],
            "price": nums[1],
            "total_excl": nums[2],
            "tax": nums[3],
            "total_incl": nums[4],
        }
        candidates.append(a)

        # Format B (seen in some PDFs): total_incl, tax, total_excl, price, qty
        b = {
            "quantity": nums[4],
            "price": nums[3],
            "total_excl": nums[2],
            "tax": nums[1],
            "total_incl": nums[0],
        }
        candidates.append(b)

        def score(c):
            q = c["quantity"]
            p = c["price"]
            te = c["total_excl"]
            t = c["tax"]
            ti = c["total_incl"]
            s = 0
            if q > 0 and q < 10000 and abs(q - round(q)) <= 0.01:
                s += 2
            if p > 0:
                s += 1
            if abs((p * q) - te) <= max(1.0, te * 0.03):
                s += 3
            if abs((te + t) - ti) <= max(1.0, ti * 0.03):
                s += 3
            return s

        best = max(candidates, key=score)
        return best if score(best) >= 4 else None

    i = 0
    while i < len(lines):
        token = lines[i]
        if ITEM_CODE_RE.match(token):
            # Gather nearby numeric tokens
            j = i + 1
            nums = []
            first_non_num_after_nums = None
            while j < len(lines) and j <= i + 16:
                if is_num(lines[j]):
                    nums.append(float(lines[j]))
                elif nums:
                    first_non_num_after_nums = j
                    break
                j += 1

            mapped = try_map_numbers(nums[:5])
            if mapped:
                desc = ""
                if first_non_num_after_nums is not None:
                    desc = lines[first_non_num_after_nums]
                elif i + 1 < len(lines):
                    # fallback if description was before numbers in weird layouts
                    desc = lines[i + 1] if not is_num(lines[i + 1]) else ""

                # Skip obvious footer noise as description
                if desc.lower().startswith(("total", "computer mania", "page:", "po date")):
                    desc = ""

                rows.append({
                    "item_number": token,
                    "description": desc,
                    "quantity": float(mapped["quantity"]),
                    "price": float(mapped["price"]),
                    "total_excl": float(mapped["total_excl"]),
                    "tax": float(mapped["tax"]),
                    "total_incl": float(mapped["total_incl"]),
                })

                i = j
                continue

        i += 1

    # Fallback for PDFs extracted as one long line (common on cloud parsers)
    if not rows:
        tokens = re.findall(r"[A-Z0-9\-/\.\"]+|\d+\.\d+|\d+", text)
        k = 0
        while k < len(tokens):
            tok = tokens[k]
            if ITEM_CODE_RE.match(tok):
                # collect next numeric values in local window
                nums = []
                nidx = []
                m = k + 1
                while m < len(tokens) and m <= k + 40 and len(nums) < 5:
                    if NUM_RE.match(tokens[m]):
                        nums.append(float(tokens[m]))
                        nidx.append(m)
                    m += 1

                mapped = try_map_numbers(nums)
                if mapped:
                    desc = ""
                    if nidx:
                        # description likely starts right after numeric block
                        ds = nidx[-1] + 1
                        de = min(ds + 10, len(tokens))
                        cand = " ".join(tokens[ds:de]).strip()
                        if cand and not cand.lower().startswith(("total", "computer", "page", "po")):
                            desc = cand

                    rows.append({
                        "item_number": tok,
                        "description": desc,
                        "quantity": float(mapped["quantity"]),
                        "price": float(mapped["price"]),
                        "total_excl": float(mapped["total_excl"]),
                        "tax": float(mapped["tax"]),
                        "total_incl": float(mapped["total_incl"]),
                    })
                    k = m
                    continue
            k += 1

    return rows


def build_excel(all_rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "PO Data"

    headers = [
        "Source PDF",
        "Purchase Order Number",
        "PO Date",
        "Item Number",
        "Description",
        "Quantity",
        "Price",
        "Total Excl",
        "Tax",
        "Total Incl",
    ]
    ws.append(headers)

    for r in all_rows:
        ws.append([
            r["source_pdf"],
            r["po_number"],
            r["po_date"],
            r["item_number"],
            r["description"],
            r["quantity"],
            r["price"],
            r["total_excl"],
            r["tax"],
            r["total_incl"],
        ])

    widths = [26, 22, 12, 16, 50, 10, 10, 12, 10, 12]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=2, min_col=6, max_col=10):
        for cell in row:
            cell.number_format = "0.00"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


files = st.file_uploader(
    "Upload one or more purchase-order PDFs",
    type=["pdf"],
    accept_multiple_files=True,
)

if st.button("Process", type="primary"):
    if not files:
        st.warning("Please upload at least one PDF.")
    else:
        all_rows = []
        warnings = []

        for f in files:
            try:
                text = extract_text_from_pdf(f)
                po_number = parse_po_number(text)
                po_date = parse_po_date(text)
                rows = parse_rows(text)

                if not rows:
                    warnings.append(f"No item rows detected in {f.name}")
                    continue

                for row in rows:
                    row["source_pdf"] = f.name
                    row["po_number"] = po_number
                    row["po_date"] = po_date
                    all_rows.append(row)
            except Exception as e:
                warnings.append(f"Failed parsing {f.name}: {e}")

        if not all_rows:
            st.error("Nothing could be parsed from uploaded PDFs.")
            if warnings:
                st.write("Warnings:")
                st.write("\n".join([f"- {w}" for w in warnings]))
        else:
            xlsx = build_excel(all_rows)
            st.success(f"Processed {len(files)} file(s), extracted {len(all_rows)} line item(s).")
            if warnings:
                with st.expander("Warnings"):
                    st.write("\n".join([f"- {w}" for w in warnings]))

            st.download_button(
                label="Download combined Excel",
                data=xlsx,
                file_name="purchase_orders_combined.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
